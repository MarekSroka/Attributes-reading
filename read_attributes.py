
# LIBRARIES
import os
from openpyxl import load_workbook
import pandas as pd

# CODE

def read_data_from_folder(folder_path):
    data = {'Object Type': [], 'Filename': [], 'Attribute Name': [], 'Attribute Value': [], 'Attribute Type': []}
    for file in os.listdir(folder_path):
        if file.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file)
            workbook = load_workbook(file_path)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                object_type_index = None
                attribute_columns = []
                for cell in worksheet[1]: 
                    if cell.value == 'objectType':
                        object_type_index = cell.column
                    elif cell.value and cell.value.startswith('attributeList.attribute.name'):
                        attribute_columns.append(cell.column)
                if object_type_index and attribute_columns:
                    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                        if row_index == 2:
                            for column_index in attribute_columns:
                                next_column_index = column_index + 1
                                if next_column_index <= worksheet.max_column:
                                    attribute_name = row[column_index - 1]
                                    data['Object Type'].append(row[object_type_index - 1])
                                    data['Filename'].append(file)  
                                    data['Attribute Name'].append(attribute_name) 
                                    if worksheet.cell(row=1, column=next_column_index).value.startswith('attributeList.attribute.valueType'):
                                        attribute_value = '<none>'                                   
                                        attribute_type = '<none>'
                                    else:
                                            next_column_value = worksheet.cell(row=1, column=next_column_index).value
                                            attribute_value = row[next_column_index - 1]
                                            if worksheet.cell(row=1, column=next_column_index).value.startswith('attributeList.attribute.real'):                                           
                                                attribute_type = 'real'
                                                attribute_value = str(worksheet.cell(row=2, column=next_column_index).value) + ' [' + str(worksheet.cell(row=2, column=next_column_index+1).value) + ']'
                                            else:
                                                attribute_type = next_column_value.split('attributeList.attribute.')[1].split('.')[0] if '.' in next_column_value else next_column_value.split('attributeList.attribute.')[1]
                                    data['Attribute Value'].append(attribute_value)
                                    data['Attribute Type'].append(attribute_type)
    return data

def main():
    folder_path = input("Enter the folder path: ")
    data = read_data_from_folder(folder_path)
    df = pd.DataFrame(data)
    
    output_file_name = os.path.basename(os.path.normpath(folder_path)) + '.xlsx'
    df.to_excel(output_file_name, index=False)  
    print(f"DataFrame saved to {output_file_name}")

if __name__ == "__main__":
    main()